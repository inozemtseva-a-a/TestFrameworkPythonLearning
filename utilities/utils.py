import inspect
import logging

import softest
from openpyxl import Workbook, load_workbook
import csv


class Utils(softest.TestCase):
    def assertListItemText(self, lists, value):
        for stop in lists:
            print("The text is: " + stop.text)
            self.soft_assert(self.assertEqual, stop.text, value)
            if stop.text == value:
                print("test passed")
            else:
                print("test failed")

        self.assert_all()

    def read_data_from_excel(file_name, sheet):
        datalist = []
        wb = load_workbook(filename=file_name)
        sh = wb[sheet]
        row_ct = sh.max_row  # how many rows are in the sheet
        col_ct = sh.max_column  # how many columns are in the sheet

        for i in range(2, row_ct + 1): #start from 2 bec. first row is headers
            row = []
            for j in range(1, col_ct + 1):
                row.append(sh.cell(row=i, column=j).value)
            datalist.append(row)
        return datalist

    def read_data_from_csv(filename):
        #create an empty list
        datalist = []

        #Open CSV file
        csvdata = open(filename,"r")

        #create CSV reader
        reader = csv.reader(csvdata)

        #skip header (1-st row)
        next(reader)

        #Add csv rows to list
        for rows in reader:
            datalist.append(rows)
        return datalist

    def custom_logger(self,log_level=logging.DEBUG):
        #Set class/method name from where it's called
        logger_name = inspect.stack()[1][3]
        #create logger
        logger = logging.getLogger(logger_name) #display the name of the file where the log is coming from and class
        logger.setLevel(log_level)
        #create console handler or file handler and set the log level
        fh = logging.FileHandler("automation.log")
        #create formatter
        formatter1 = logging.Formatter(fmt="%(asctime)s - %(levelname)s : %(message)s", datefmt="%m/%d/%Y %I:%M:%S %p")
        #add formatter to console or file handler
        fh.setFormatter(formatter1)
        #add console handler to logger
        logger.addHandler(fh)
        return logger